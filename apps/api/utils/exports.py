import logging
import os
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

logger = logging.getLogger(__name__)


def generate_pdf_report(title: str, data: dict[str, Any], output_path: str) -> str:
    """
    Generate a basic PDF report using ReportLab.

    Args:
        title: The report title displayed at the top of the document.
        data: A dictionary of section_name -> value pairs to include in the report.
        output_path: Full filesystem path where the PDF will be saved.

    Returns:
        The output_path of the generated file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()
    elements: list[Any] = []

    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.3 * inch))

    heading_style = styles["Heading2"]

    for section_name, section_data in data.items():
        elements.append(Paragraph(section_name, heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        if isinstance(section_data, dict):
            table_data = [["Key", "Value"]]
            for key, value in section_data.items():
                table_data.append([str(key), str(value)])

            table = Table(table_data, colWidths=[3 * inch, 4 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B7BC4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, 0), 11),
                        ("FONTSIZE", (0, 1), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F4FD")]),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(table)
        elif isinstance(section_data, list):
            for item in section_data:
                elements.append(Paragraph(f"• {item}", styles["BodyText"]))
        else:
            elements.append(Paragraph(str(section_data), styles["BodyText"]))

        elements.append(Spacer(1, 0.2 * inch))

    doc.build(elements)
    logger.info("PDF report generated: %s", output_path)
    return output_path


def generate_excel_report(title: str, data: list[dict[str, Any]], output_path: str) -> str:
    """
    Generate an Excel workbook using openpyxl.

    Args:
        title: Used as the worksheet name.
        data: A list of dictionaries. Dictionary keys become column headers.
        output_path: Full filesystem path where the Excel file will be saved.

    Returns:
        The output_path of the generated file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet names are limited to 31 characters

    if not data:
        ws.cell(row=1, column=1, value="No data available")
        wb.save(output_path)
        logger.info("Excel report generated (empty): %s", output_path)
        return output_path

    headers = list(data[0].keys())

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_length + 2, 50)

    wb.save(output_path)
    logger.info("Excel report generated: %s", output_path)
    return output_path
