import asyncio
import logging
import os
from datetime import datetime, timedelta, timezone

from celery import shared_task
from sqlalchemy import func, select

from core.database import async_session
from models.deliverable import Deliverable
from models.enums import AccountStatus, DeliverableStatus, TaskStatus
from models.subscription import Subscription
from models.task import Task
from models.user import User

logger = logging.getLogger(__name__)

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")


def _ensure_reports_dir() -> str:
    abs_path = os.path.abspath(REPORTS_DIR)
    os.makedirs(abs_path, exist_ok=True)
    return abs_path


@shared_task(name="generate_weekly_report")
def generate_weekly_report() -> str:
    from utils.exports import generate_pdf_report

    reports_dir = _ensure_reports_dir()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(reports_dir, f"weekly_report_{timestamp}.pdf")

    logger.info("Starting weekly report generation...")

    data = asyncio.run(_fetch_weekly_data())

    result_path = generate_pdf_report(
        title="Creo Weekly Report",
        data=data,
        output_path=output_path,
    )

    logger.info("Weekly report generated successfully: %s", result_path)
    return result_path


async def _fetch_weekly_data() -> dict:
    now = datetime.now(timezone.utc)
    week_ago = now - timedelta(days=7)

    async with async_session() as db:
        active_tasks_result = await db.execute(
            select(func.count(Task.id)).where(
                Task.status.in_([TaskStatus.pending, TaskStatus.in_progress])
            )
        )
        active_tasks = active_tasks_result.scalar() or 0

        completed_result = await db.execute(
            select(func.count(Task.id)).where(
                Task.status == TaskStatus.approved,
                Task.updated_at >= week_ago,
            )
        )
        completed_this_week = completed_result.scalar() or 0

        pending_review_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.status == DeliverableStatus.pending_approval
            )
        )
        pending_review = pending_review_result.scalar() or 0

        overdue_result = await db.execute(
            select(func.count(Task.id)).where(Task.status == TaskStatus.overdue)
        )
        overdue = overdue_result.scalar() or 0

        submitted_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.created_at >= week_ago
            )
        )
        total_submitted = submitted_result.scalar() or 0

        approved_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.status == DeliverableStatus.approved,
                Deliverable.approved_at >= week_ago,
            )
        )
        approved = approved_result.scalar() or 0

        rejected_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.status == DeliverableStatus.rejected,
                Deliverable.rejected_at >= week_ago,
            )
        )
        rejected = rejected_result.scalar() or 0

    return {
        "Report Period": {
            "Generated At": now.isoformat(),
            "Report Type": "Weekly Operations Summary",
        },
        "Active Tasks": {
            "Total Tasks": active_tasks,
            "Completed This Week": completed_this_week,
            "Pending Review": pending_review,
            "Overdue": overdue,
        },
        "Deliverables": {
            "Total Submitted": total_submitted,
            "Approved": approved,
            "Rejected": rejected,
            "Pending Approval": pending_review,
        },
    }


@shared_task(name="generate_monthly_report")
def generate_monthly_report() -> str:
    from utils.exports import generate_pdf_report

    reports_dir = _ensure_reports_dir()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(reports_dir, f"monthly_report_{timestamp}.pdf")

    logger.info("Starting monthly report generation...")

    data = asyncio.run(_fetch_monthly_data())

    result_path = generate_pdf_report(
        title="Creo Monthly Report",
        data=data,
        output_path=output_path,
    )

    logger.info("Monthly report generated successfully: %s", result_path)
    return result_path


async def _fetch_monthly_data() -> dict:
    now = datetime.now(timezone.utc)
    month_ago = now - timedelta(days=30)

    async with async_session() as db:
        new_clients_result = await db.execute(
            select(func.count(User.id)).where(
                User.role == "client",
                User.created_at >= month_ago,
                User.deleted_at.is_(None),
            )
        )
        new_clients = new_clients_result.scalar() or 0

        active_clients_result = await db.execute(
            select(func.count(User.id)).where(
                User.role == "client",
                User.account_status == AccountStatus.active,
                User.deleted_at.is_(None),
            )
        )
        total_active = active_clients_result.scalar() or 0

        churned_result = await db.execute(
            select(func.count(User.id)).where(
                User.role == "client",
                User.account_status == AccountStatus.lapsed,
                User.updated_at >= month_ago,
                User.deleted_at.is_(None),
            )
        )
        churned = churned_result.scalar() or 0

        sla_breaches_result = await db.execute(
            select(func.count(Task.id)).where(
                Task.status == TaskStatus.overdue,
                Task.updated_at >= month_ago,
            )
        )
        sla_breaches = sla_breaches_result.scalar() or 0

        total_deliverables_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.created_at >= month_ago
            )
        )
        total_deliverables = total_deliverables_result.scalar() or 0

        on_time_result = await db.execute(
            select(func.count(Deliverable.id)).where(
                Deliverable.status == DeliverableStatus.approved,
                Deliverable.created_at >= month_ago,
            )
        )
        on_time = on_time_result.scalar() or 0

    return {
        "Report Period": {
            "Generated At": now.isoformat(),
            "Report Type": "Monthly Business Summary",
        },
        "Client Growth": {
            "New Clients This Month": new_clients,
            "Total Active Clients": total_active,
            "Churned Clients": churned,
            "Net Growth": new_clients - churned,
        },
        "SLA Performance": {
            "Total Deliverables": total_deliverables,
            "Delivered On Time": on_time,
            "SLA Breaches": sla_breaches,
            "SLA Compliance Rate": (
                f"{(on_time / total_deliverables * 100):.1f}%"
                if total_deliverables > 0
                else "N/A"
            ),
        },
    }


@shared_task(name="generate_financial_report")
def generate_financial_report() -> str:
    reports_dir = _ensure_reports_dir()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(reports_dir, f"financial_report_{timestamp}.xlsx")

    logger.info("Starting financial report generation...")

    summary_data, revenue_by_plan = asyncio.run(_fetch_financial_data())

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = list(summary_data[0].keys()) if summary_data else ["Metric", "Value"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B7BC4", end_color="2B7BC4", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_data in enumerate(summary_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    ws_plan = wb.create_sheet("Revenue by Plan")
    plan_headers = list(revenue_by_plan[0].keys()) if revenue_by_plan else ["Plan", "Subscribers", "Monthly Price", "Total Revenue"]

    for col_idx, header in enumerate(plan_headers, start=1):
        cell = ws_plan.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_data in enumerate(revenue_by_plan, start=2):
        for col_idx, header in enumerate(plan_headers, start=1):
            ws_plan.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    wb.save(output_path)
    logger.info("Financial report generated successfully: %s", output_path)
    return output_path


async def _fetch_financial_data() -> tuple[list[dict], list[dict]]:
    from models.plan import Plan

    now = datetime.now(timezone.utc)

    async with async_session() as db:
        active_subs_result = await db.execute(
            select(func.count(Subscription.id)).where(
                Subscription.status == "active"
            )
        )
        active_subs = active_subs_result.scalar() or 0

        mrr_result = await db.execute(
            select(func.sum(Plan.monthly_price))
            .join(Subscription, Subscription.plan_id == Plan.id)
            .where(Subscription.status == "active")
        )
        mrr = mrr_result.scalar() or 0.0

        arpu = (mrr / active_subs) if active_subs > 0 else 0.0

        plan_breakdown_result = await db.execute(
            select(
                Plan.name,
                Plan.monthly_price,
                func.count(Subscription.id).label("subscribers"),
            )
            .join(Subscription, Subscription.plan_id == Plan.id)
            .where(Subscription.status == "active")
            .group_by(Plan.id, Plan.name, Plan.monthly_price)
        )
        plan_rows = plan_breakdown_result.all()

    summary_data = [
        {"Metric": "Total MRR (INR)", "Value": round(mrr, 2)},
        {"Metric": "Active Subscriptions", "Value": active_subs},
        {"Metric": "Average Revenue Per User", "Value": round(arpu, 2)},
        {"Metric": "Report Generated", "Value": now.isoformat()},
    ]

    revenue_by_plan = []
    for row in plan_rows:
        subscribers = row.subscribers
        price = float(row.monthly_price) if row.monthly_price else 0.0
        revenue_by_plan.append({
            "Plan": row.name,
            "Subscribers": subscribers,
            "Monthly Price": price,
            "Total Revenue": round(price * subscribers, 2),
        })

    return summary_data, revenue_by_plan
